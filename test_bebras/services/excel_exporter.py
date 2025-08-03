import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.utils import timezone
import pytz 

from ..models import Attempt, Group, Test, GroupMetadata, TestAssignment
from django.contrib.auth import get_user_model

def generate_attempts_xlsx_report(user, group_id_filter=None, test_id_filter=None):

    if not user.is_staff:
        test_assignments_queryset = TestAssignment.objects.filter(
            assigned_by=user,
            group__groupmetadata__created_by=user,
            test__creator=user
        )
    else:
        test_assignments_queryset = TestAssignment.objects.all()

    filtered_group_name = None
    if group_id_filter:
        test_assignments_queryset = test_assignments_queryset.filter(group__id=group_id_filter)
    if test_id_filter:
        test_assignments_queryset = test_assignments_queryset.filter(test__id=test_id_filter)
    
    
    if not test_assignments_queryset.exists():
        return None
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Resultados de Intentos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    headers = [
        "N° Intento", "Estudiante", "Correo Electrónico", "Grupo(s)", "Nombre del Test",
        "Fecha de Intento", "Puntuación", "Respuestas Correctas",
        "Límite de Tiempo (min)", "Permite retroceder", "Permite no responder", 'Fecha Finalizado'
    ]
    sheet.append(headers)

    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_aligned_text
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20

    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 15

    row_num = 2

    current_user_id = None
    current_test_id = None
    attempt_counter = 0

    chile_tz = None
    if settings.USE_TZ and hasattr(settings, 'TIME_ZONE'):
        try:
            chile_tz = pytz.timezone(settings.TIME_ZONE)
        except pytz.exceptions.UnknownTimeZoneError:
            chile_tz = None

    for test_assignment in test_assignments_queryset:
        attempts_for_assignment = Attempt.objects.filter(
            test=test_assignment.test,
            user__groups=test_assignment.group
        ).order_by('user__username', 'date_taken')
        
        current_user_id = None
        attempt_counter = 0

        for attempt in attempts_for_assignment:
            if attempt.user.id != current_user_id:
                attempt_counter = 1
                current_user_id = attempt.user.id
            else:
                attempt_counter += 1

            groups_display = test_assignment.group.name

            local_date_taken = attempt.date_taken
            if chile_tz and timezone.is_aware(local_date_taken):
                local_date_taken = local_date_taken.astimezone(chile_tz)
            elif chile_tz and timezone.is_naive(local_date_taken):
                local_date_taken = timezone.make_aware(local_date_taken, timezone.utc).astimezone(chile_tz)

            sheet.cell(row=row_num, column=1, value=attempt_counter)
            sheet.cell(row=row_num, column=2, value=attempt.user.get_full_name() or attempt.user.username)
            sheet.cell(row=row_num, column=3, value=attempt.user.email)
            sheet.cell(row=row_num, column=4, value=groups_display)
            sheet.cell(row=row_num, column=5, value=attempt.test.name)
            sheet.cell(row=row_num, column=6, value=attempt.date_taken.strftime("%Y-%m-%d %H:%M:%S"))
            sheet.cell(row=row_num, column=7, value=float(attempt.score))
            sheet.cell(row=row_num, column=8, value=attempt.correct_count)
            sheet.cell(row=row_num, column=9, value=attempt.test.maximum_time.total_seconds() / 60 if attempt.test.maximum_time else "N/A")
            sheet.cell(row=row_num, column=10, value="Sí" if attempt.test.allow_backtracking else "No")
            sheet.cell(row=row_num, column=11, value="Sí" if attempt.test.allow_no_response else "No")
            sheet.cell(row=row_num, column=12, value=attempt.end_time.strftime("%Y-%m-%d %H:%M:%S") if attempt.end_time else "-")

            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=row_num, column=col_idx).border = border_style

            row_num += 1

    from io import BytesIO
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0) 

    return excel_file